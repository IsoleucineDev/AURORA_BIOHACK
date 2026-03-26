"""
Lector Serial ESP32 → Excel  |  Sensores GSR + MAX30105 (IR)
Llena automáticamente Datos_GSR_IR_Sensor.xlsx

Sesiones en orden:
  1. Reposo      - 1  minuto
  2. Reposo      - 5  minutos
  3. Reposo      - 10 minutos
  4. Movimiento  - 1  minuto
  5. Movimiento  - 5  minutos
  6. Movimiento  - 10 minutos

Uso:
  pip install pyserial openpyxl
  python lector_gsr_ir.py
"""

import serial
import serial.tools.list_ports
import openpyxl
import time
import sys
import os
import threading

EXCEL_FILE = "Datos_GSR_IR_Sensor.xlsx"

SESIONES = [
    ("Reposo_1min",      1,   False, "REPOSO      — 1 minuto"),
    ("Reposo_5min",      5,   False, "REPOSO      — 5 minutos"),
    ("Reposo_10min",     10,  False, "REPOSO      — 10 minutos"),
    ("Movimiento_1min",  1,   True,  "MOVIMIENTO  — 1 minuto"),
    ("Movimiento_5min",  5,   True,  "MOVIMIENTO  — 5 minutos"),
    ("Movimiento_10min", 10,  True,  "MOVIMIENTO  — 10 minutos"),
]


def listar_puertos():
    puertos = serial.tools.list_ports.comports()
    if not puertos:
        print("❌ No se encontraron puertos seriales.")
        sys.exit(1)
    print("\nPuertos disponibles:")
    for i, p in enumerate(puertos):
        print(f"  [{i}] {p.device}  —  {p.description}")
    return puertos


def seleccionar_puerto(puertos):
    if len(puertos) == 1:
        print(f"\n✅ Puerto detectado automáticamente: {puertos[0].device}")
        return puertos[0].device
    idx = input("\nElige el número del puerto del ESP32: ").strip()
    return puertos[int(idx)].device


def cargar_excel():
    if not os.path.exists(EXCEL_FILE):
        print(f"\n❌ No se encontró '{EXCEL_FILE}' en esta carpeta.")
        print("   Asegúrate de tener el archivo Excel en el mismo directorio que este script.")
        sys.exit(1)
    wb = openpyxl.load_workbook(EXCEL_FILE)
    print(f"✅ Excel cargado: {EXCEL_FILE}")
    return wb


def parsear_linea(linea):
    """
    Parsea líneas con formato: 'GSR:1234, IR:56789'
    Retorna (gsr, ir) como enteros, o None si falla.
    """
    try:
        partes = linea.split(",")
        gsr_part = partes[0].strip()
        ir_part  = partes[1].strip()
        gsr = int(gsr_part.split(":")[1].strip())
        ir  = int(ir_part.split(":")[1].strip())
        return gsr, ir
    except Exception:
        return None


def barra_progreso(actual, total, ancho=35):
    llenos = int(ancho * actual / total)
    barra = "█" * llenos + "░" * (ancho - llenos)
    pct = 100 * actual / total
    return f"[{barra}] {pct:5.1f}%  ({actual}/{total}s)"


def ejecutar_sesion(ser, wb, idx_sesion):
    sheet_name, duracion_min, es_movimiento, etiqueta = SESIONES[idx_sesion]
    total_seg = duracion_min * 60
    tipo_str  = "🏃 MOVIMIENTO" if es_movimiento else "🧘 REPOSO"

    print(f"\n{'═'*58}")
    print(f"  Sesión {idx_sesion + 1}/6  |  {tipo_str}")
    print(f"  {etiqueta}")
    print(f"  Duración: {duracion_min} minuto{'s' if duracion_min > 1 else ''}  ({total_seg} muestras)")
    print(f"{'═'*58}")

    if es_movimiento:
        print("\n  ⚠️  Esta sesión es en MOVIMIENTO.")
        print("     Prepárate para moverte cuando empiece la cuenta.\n")

    print("  Presiona ENTER para iniciar esta sesión...")
    input()

    ws = wb[sheet_name]
    ser.reset_input_buffer()

    datos      = []       # lista de (gsr, ir) por segundo
    buffer_raw = []       # buffer para promediar dentro de cada segundo
    segundo_actual = 0
    t_inicio   = time.time()
    t_ultimo_seg = t_inicio

    print(f"\n  📡 Midiendo... {'¡Empieza a moverte!' if es_movimiento else 'Mantente en reposo.'}\n")

    try:
        while segundo_actual < total_seg:
            ahora = time.time()

            # Leer líneas disponibles del serial
            while ser.in_waiting > 0:
                try:
                    linea = ser.readline().decode("utf-8", errors="ignore").strip()
                    parsed = parsear_linea(linea)
                    if parsed:
                        buffer_raw.append(parsed)
                except Exception:
                    pass

            # Cada segundo: promediar buffer y guardar
            if ahora - t_ultimo_seg >= 1.0:
                segundo_actual += 1
                t_ultimo_seg = ahora

                if buffer_raw:
                    gsr_prom = sum(x[0] for x in buffer_raw) / len(buffer_raw)
                    ir_prom  = sum(x[1] for x in buffer_raw) / len(buffer_raw)
                    buffer_raw.clear()
                else:
                    # Si no hubo dato ese segundo, repetir último o poner 0
                    gsr_prom = datos[-1][0] if datos else 0
                    ir_prom  = datos[-1][1] if datos else 0

                datos.append((round(gsr_prom), round(ir_prom)))

                # Escribir en Excel (fila 4 = segundo 1)
                fila = segundo_actual + 3
                ws.cell(row=fila, column=3).value = round(gsr_prom)
                ws.cell(row=fila, column=4).value = round(ir_prom)

                # Progreso en consola
                barra = barra_progreso(segundo_actual, total_seg)
                mm = (segundo_actual - 1) // 60
                ss = (segundo_actual - 1) % 60
                print(f"  {barra}  |  GSR: {round(gsr_prom):5d}  IR: {round(ir_prom):7d}  [{mm:02d}:{ss:02d}]", end="\r")

            time.sleep(0.01)

    except KeyboardInterrupt:
        print("\n\n  ⏹️  Sesión interrumpida por el usuario.")

    wb.save(EXCEL_FILE)

    # Resumen
    if datos:
        gsrs = [d[0] for d in datos]
        irs  = [d[1] for d in datos]
        print(f"\n\n  ✅ Sesión completada y guardada  →  {sheet_name}")
        print(f"     Muestras: {len(datos)}s  |  "
              f"GSR prom: {sum(gsrs)/len(gsrs):.0f}  |  "
              f"IR prom: {sum(irs)/len(irs):.0f}")
    else:
        print(f"\n\n  ⚠️  No se recibieron datos para {sheet_name}.")


def main():
    print("\n╔══════════════════════════════════════════════════════════╗")
    print("║   ESP32 → Excel  |  GSR + MAX30105 (IR)                 ║")
    print("║   6 sesiones: Reposo (1/5/10 min) + Movimiento (ídem)   ║")
    print("╚══════════════════════════════════════════════════════════╝\n")

    puertos = listar_puertos()
    puerto  = seleccionar_puerto(puertos)

    try:
        ser = serial.Serial(puerto, 115200, timeout=1)
        print(f"✅ Conectado a {puerto} @ 115200 baud\n")
    except serial.SerialException as e:
        print(f"❌ Error al abrir puerto: {e}")
        sys.exit(1)

    wb = cargar_excel()

    print("\nSe realizarán 6 sesiones en este orden:")
    for i, (_, dur, mov, etiqueta) in enumerate(SESIONES, 1):
        tipo = "🏃 Movimiento" if mov else "🧘 Reposo"
        print(f"  {i}. {tipo:<20}  {dur:>2} min  ({dur*60}s)")

    print("\nPuedes saltar cualquier sesión con Ctrl+C y continuar a la siguiente.")
    print("El Excel se guarda automáticamente al terminar cada sesión.\n")

    for i in range(len(SESIONES)):
        try:
            ejecutar_sesion(ser, wb, i)
        except KeyboardInterrupt:
            print(f"\n  ⏩ Sesión saltada.")

        if i < len(SESIONES) - 1:
            resp = input("\n  ¿Continuar con la siguiente sesión? (Enter = sí / 'n' = salir): ").strip().lower()
            if resp == "n":
                print("\n⏹️  Sesión terminada por el usuario.")
                break

    ser.close()
    print(f"\n🎉 Listo. Todos los datos guardados en '{EXCEL_FILE}'")


if __name__ == "__main__":
    main()
